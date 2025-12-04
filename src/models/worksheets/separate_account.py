from __future__ import annotations

from typing import List, Optional, Sequence

from openpyxl.styles import Alignment, Border, Font, Side

from .base import BaseWorksheet


class SeparateAccountsWorksheet(BaseWorksheet):
    """Worksheet for separate accounts with predefined headers and formatting."""

    def __init__(self, workbook_path: str) -> None:
        self.headers: List[str] = ["日期", "摘要", "部門", "借方金額", "貸方金額", "承辦律師"]
        super().__init__(workbook_path, "程式RUN後檔案")

    def format_worksheet(self) -> None:
        """格式化工作表，設定欄寬和表頭。"""
        letters = ["A", "B", "C", "D", "E", "F"]
        # 設定欄寬
        width_offset = 2
        self.ws.column_dimensions["A"].width = 15.9 + width_offset
        self.ws.column_dimensions["B"].width = 50.5 + width_offset
        self.ws.column_dimensions["C"].width = 13 + width_offset
        self.ws.column_dimensions["D"].width = 15.9 + width_offset
        self.ws.column_dimensions["E"].width = 15.9 + width_offset
        self.ws.column_dimensions["F"].width = 15.9 + width_offset

        # 建立欄位
        for col, letter in enumerate(letters, start=1):
            self.ws.merge_cells(f"{letter}1:{letter}2")
            cell = self.ws.cell(row=1, column=col)

            # 設定 Cell
            cell.value = self.headers[col - 1]  # 給值
            cell.font = Font(name="微軟正黑體", size=13)
            for row in self.ws[f"{letters[0]}1:{letters[-1]}2"]:
                for single_cell in row:
                    single_cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_data_to_worksheet(
        self,
        data_rows: Sequence[Sequence[object]],
        *,
        total_debit: Optional[int] = None,
        total_credit: Optional[int] = None,
    ) -> None:
        """將資料寫入工作表。"""
        first_empty_row = self.ws.max_row + 1
        # Note: In GUI mode, we don't prompt for overwrite here
        # The overwrite check is handled by QtSeparateAccountsWorksheet

        # 開始寫入資料到工作表
        for row_index, row in enumerate(data_rows):
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=first_empty_row + row_index, column=column_index + 1)
                cell.value = row[column_index]

                cell.font = Font(name="Courier New", size=13)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if column_index == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="bottom", wrap_text=True)

        if total_credit is not None or total_debit is not None:
            self.ws.append(["", "合計", "", total_debit, total_credit, ""])
            total_row = self.ws.max_row
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=total_row, column=column_index + 1)

                cell.font = Font(name="Courier New", size=13)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if column_index == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        self.save(self.workbook_path)
