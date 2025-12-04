from __future__ import annotations

import os
from abc import ABC, abstractmethod
from typing import Optional, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class BaseWorksheet(ABC):
    """Base class for Excel worksheet operations."""

    def __init__(self, workbook_path: str, sheet_title: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_title = sheet_title

        if not os.path.exists(workbook_path):
            self.wb = Workbook()
            self.is_new_workbook = True
        else:
            self.is_new_workbook = False
            self.wb = load_workbook(workbook_path)

        self.ws: Worksheet
        self.init_ws()

    def init_ws(self) -> None:
        """Initialize worksheet. Should be overridden by subclasses if needed."""
        # 如果是新建 Workbook 而有預設建立的工作表，刪除它
        if self.is_new_workbook and "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # 如果目標工作表存在則使用它，否則創建一個新的工作表
        if self.sheet_title in self.wb.sheetnames:
            # Use existing worksheet
            self.ws = self.wb[self.sheet_title]
        else:
            # Create new worksheet
            self.create_worksheet()

    def create_worksheet(self) -> None:
        """Create and format a new worksheet."""
        self.ws = self.wb.create_sheet(self.sheet_title)
        self.format_worksheet()
        self.save(self.workbook_path)

    def save(self, filename: str) -> None:
        """Save the workbook to a file."""
        try:
            self.wb.save(filename)
        except PermissionError as err:
            print(
                f"無法儲存變更到檔案名稱{filename}",
                f"請確認目標檔案'{filename}'是否已經關閉，若透過其他程式開啟該檔案，則無法在此程式更改該檔案！",
            )
            print(err)

    @abstractmethod
    def format_worksheet(self) -> None:
        """格式化工作表外觀。"""

    @abstractmethod
    def write_data_to_worksheet(self, data_rows: Sequence[Sequence[object]], **kwargs: object) -> None:
        """將資料寫入工作表。"""
