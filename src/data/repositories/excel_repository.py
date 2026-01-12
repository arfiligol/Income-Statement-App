from __future__ import annotations

import collections.abc
import unicodedata
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

class ExcelRepository:
    """Repository for low-level Excel operations (read/write)."""

    WORKFLOW_TARGET_SHEET = "明細分類帳"

    def get_read_engine(self, file_path: str) -> str:
        """Determines the engine for pandas read_excel."""
        return "xlrd" if file_path.lower().endswith(".xls") else "openpyxl"

    def read_dataframe(self, file_path: str) -> tuple[pd.DataFrame, str]:
        """
        Reads the target sheet from Excel file into a DataFrame.
        Returns: (DataFrame, actual_sheet_name)
        """
        engine = self.get_read_engine(file_path)
        try:
            xls = pd.ExcelFile(file_path, engine=engine)
        except Exception as e:
            raise IOError(f"無法開啟檔案: {e}") from e

        sheet_name = self._resolve_target_sheet_name(xls.sheet_names)
        df = xls.parse(sheet_name=sheet_name)
        if not isinstance(df, pd.DataFrame):
             raise ValueError("解析工作表失敗")
        return df, sheet_name

    def load_workbook(self, file_path: str) -> Workbook:
        """Loads an openpyxl Workbook."""
        return load_workbook(file_path)

    def save_workbook(self, workbook: Workbook, file_path: str) -> None:
        """Saves the workbook."""
        workbook.save(file_path)

    def _resolve_target_sheet_name(self, sheet_names: collections.abc.Sequence[str]) -> str:
        """Fuzzy matches the target sheet name."""
        target = self.WORKFLOW_TARGET_SHEET
        if target in sheet_names:
            return target

        normalized_target = self._normalize_text(target)

        def normalized(name: str) -> str:
            return self._normalize_text(name)

        for name in sheet_names:
            if normalized(name) == normalized_target:
                return name

        for name in sheet_names:
            norm = normalized(name)
            if normalized_target in norm or norm in normalized_target:
                return name

        available = "\n".join(f"- {name}" for name in sheet_names) or "<無任何工作表>"
        raise ValueError(
            f"找不到預期的律師備註工作表 '{target}'。\n"
            f"目前偵測到的工作表：\n{available}"
        )

    def _normalize_text(self, value: Any) -> str:
        text = unicodedata.normalize("NFKC", str(value)).strip()
        text = text.replace(" ", "").replace("\u3000", "")
        text = "".join(ch for ch in text if not ch.isspace())
        return text.lower()
