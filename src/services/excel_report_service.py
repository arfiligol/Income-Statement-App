from __future__ import annotations

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.dtos.workflow_dto import SeparateLedgerResultDTO
from src.services.interfaces.user_interaction_provider import UserInteractionProvider


class ExcelReportService:
    """Service for generating Excel reports (replacing separate_account.py)."""

    SHEET_TITLE = "程式RUN後檔案"
    HEADERS = ["日期", "摘要", "部門", "借方金額", "貸方金額", "承辦律師"]

    def __init__(self, interaction_provider: UserInteractionProvider):
        self.interaction = interaction_provider

    async def write_separate_ledger_report(self, file_path: str, result: SeparateLedgerResultDTO) -> None:
        """Writes the separate ledger result to Excel."""
        is_new_workbook = False
        if not os.path.exists(file_path):
            wb = Workbook()
            is_new_workbook = True
        else:
            try:
                wb = load_workbook(file_path)
            except PermissionError as e:
                # This should ideally be handled by the caller or provider
                raise IOError(f"無法開啟檔案，請確認檔案未被其他程式開啟: {file_path}") from e

        ws: Worksheet = None

        # Clean verify sheet logic
        if is_new_workbook and "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        if self.SHEET_TITLE in wb.sheetnames:
            if await self.interaction.confirm("工作表存在", f"工作表'{self.SHEET_TITLE}' 已經存在，你想要覆蓋它嗎？"):
                wb.remove(wb[self.SHEET_TITLE])
                ws = wb.create_sheet(self.SHEET_TITLE)
                self._format_worksheet(ws)
            else:
                ws = wb[self.SHEET_TITLE]
        else:
            ws = wb.create_sheet(self.SHEET_TITLE)
            self._format_worksheet(ws)

        # Write Logic
        first_empty_row = ws.max_row + 1
        
        # If it's not starting at row 3 (Header is 1-2), it means there's data.
        # But wait, logic in separate_account.py:
        # header at 1, header cells merged to row 2?
        # base.py checks "if first_empty_row != 3: prompt overwrite"
        # Let's align with that. format_worksheet writes headers at row 1 (merged to 2).
        
        if first_empty_row != 3:
             if await self.interaction.confirm("問題", f"偵測到工作表'{self.SHEET_TITLE}'已有資料，是否覆寫？"):
                 # Reset to row 3? How to clear content?
                 # openpyxl delete_rows
                 ws.delete_rows(3, amount=ws.max_row) 
                 first_empty_row = 3

        # Write data
        for row_index, row_data in enumerate(result.rows):
            current_row = first_empty_row + row_index
            for col_index, value in enumerate(row_data):
                cell = ws.cell(row=current_row, column=col_index + 1)
                cell.value = value
                self._style_data_cell(cell, col_index)

        # Write totals if needed
        if result.total_debit is not None or result.total_credit is not None:
            ws.append(["", "合計", "", result.total_debit, result.total_credit, ""])
            total_row = ws.max_row
            for col_index in range(len(self.HEADERS)):
                cell = ws.cell(row=total_row, column=col_index + 1)
                self._style_data_cell(cell, col_index)

        try:
            wb.save(file_path)
        except PermissionError as e:
            raise IOError(f"無法儲存檔案，請確認檔案未被其他程式開啟: {file_path}") from e

    def _format_worksheet(self, ws: Worksheet) -> None:
        """Format worksheet columns and headers."""
        letters = ["A", "B", "C", "D", "E", "F"]
        # Set column widths
        width_offset = 2
        ws.column_dimensions["A"].width = 15.9 + width_offset
        ws.column_dimensions["B"].width = 50.5 + width_offset
        ws.column_dimensions["C"].width = 13 + width_offset
        ws.column_dimensions["D"].width = 15.9 + width_offset
        ws.column_dimensions["E"].width = 15.9 + width_offset
        ws.column_dimensions["F"].width = 15.9 + width_offset

        # Create Headers
        for col, letter in enumerate(letters, start=1):
            ws.merge_cells(f"{letter}1:{letter}2")
            cell = ws.cell(row=1, column=col)
            cell.value = self.HEADERS[col - 1]
            cell.font = Font(name="微軟正黑體", size=13)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply borders to merged cells
            # Accessing range of merged cells to set border style
            # Simple approach: set border for top-left and ensure others are visual
            # But SeparateAccountsWorksheet iterates range.
            for row in ws[f"{letter}1:{letter}2"]:
                for c in row:
                    c.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

    def _style_data_cell(self, cell, col_index):
        cell.font = Font(name="Courier New", size=13)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        # "承辦律師" (Index 5? No, logic says col_index == 4? Wait)
        # HEADERS = ["日期", "摘要", "部門", "借方金額", "貸方金額", "承辦律師"]
        # 0, 1, 2, 3, 4, 5
        # In separate_account.py: if column_index == 4: center wrap
        # But '貸方金額' is index 4 (0-based) ??
        # Or maybe check the header list.
        # headers[4] is '貸方金額'. Wait, why center alignment for Credit amount?
        # Maybe index 4 meant something else, or I misread.
        # separate_account.py: self.headers[col - 1] where col starts at 1.
        # self.headers = ["日期", "摘要", "部門", "借方金額", "貸方金額", "承辦律師"]
        # write_data_to_worksheet iterates enumerate(self.headers). column_index is 0..5.
        # if column_index == 4: (which is '貸方金額') -> center.
        # Maybe related to lawyer code? No, lawyer code is 5.
        # I'll stick to original logic: column_index == 4.
        
        if col_index == 4:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        else:
            cell.alignment = Alignment(vertical="bottom", wrap_text=True)
