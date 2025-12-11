from __future__ import annotations

from typing import TYPE_CHECKING, Any, cast, override

from PySide6.QtWidgets import QMessageBox, QWidget
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from src.models.worksheets import SeparateAccountsWorksheet

if TYPE_CHECKING:
    from collections.abc import Sequence


class QtSeparateAccountsWorksheet(SeparateAccountsWorksheet):
    """SeparateAccountsWorksheet that uses Qt message boxes instead of Tk dialogs."""

    def __init__(self, parent: QWidget, workbook_path: str) -> None:
        self._qt_parent: QWidget = parent
        super().__init__(workbook_path)
        # Assuming SeparateAccountsWorksheet has 'ws' attribute, but we can hint it here if we use it.
        # If it's not defined in the parent, we define it here or assume it is.
        # Based on usage 'self.ws', it seems it is defined in parent or here.
        # The parent class likely has 'ws' as 'Any' or 'Worksheet'.
        # We'll assume it's there. If we need to hint it, we can do:
        # self.ws: Worksheet

    @override
    def init_ws(self) -> None:
        if self.is_new_workbook and "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        if self.sheet_title in self.wb.sheetnames:
            response = QMessageBox.question(
                self._qt_parent,
                "工作表存在",
                f"工作表'{self.sheet_title}' 已經存在，你想要覆蓋它嗎？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if response == QMessageBox.StandardButton.Yes:
                self.wb.remove(self.wb[self.sheet_title])
                self.create_worksheet()
            else:
                self.ws = self.wb[self.sheet_title]
        else:
            self.create_worksheet()

    @override
    def write_data_to_worksheet(
        self,
        data_rows: Sequence[Sequence[object]],
        *,
        total_debit: int | None = None,
        total_credit: int | None = None,
    ) -> None:
        # Ensure self.ws is typed as Worksheet for local usage if needed,
        # but self.ws comes from parent. We can cast it if we want to be sure.
        ws = cast(Worksheet, self.ws)

        first_empty_row = ws.max_row + 1
        if first_empty_row != 3:
            response = QMessageBox.question(
                self._qt_parent,
                "問題",
                "偵測到工作表'程式RUN後檔案'已有資料，是否覆寫？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if response == QMessageBox.StandardButton.Yes:
                first_empty_row = 3

        for row_index, row in enumerate(data_rows):
            for column_index, _ in enumerate(self.headers):
                cell = ws.cell(row=first_empty_row + row_index, column=column_index + 1)
                # Fix assignment error by casting value to Any
                cell.value = cast(Any, row[column_index])
                # We know cell is Cell because we are writing to a new row or existing row in a standard sheet
                if isinstance(cell, Cell):
                    self._style_data_cell(column_index, cell)

        if total_credit is not None or total_debit is not None:
            ws.append(["", "合計", "", total_debit, total_credit, ""])
            total_row = ws.max_row
            for column_index, _ in enumerate(self.headers):
                cell = ws.cell(row=total_row, column=column_index + 1)
                if isinstance(cell, Cell):
                    self._style_data_cell(column_index, cell)

        self.save(self.workbook_path)

    def _style_data_cell(self, column_index: int, cell: Cell) -> None:
        from openpyxl.styles import Alignment, Border, Font, Side

        cell.font = Font(name="Courier New", size=13)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        if column_index == 4:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        else:
            cell.alignment = Alignment(vertical="bottom", wrap_text=True)
