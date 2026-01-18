import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

from application.ports.gateways import ReportGateway
from common.errors import InfrastructureError
from common.types import Result
from domain.dto.separate_ledger import SeparateLedgerResult


class ExcelReportGateway(ReportGateway):
    """
    Implementation of ReportGateway using openpyxl.
    """

    def generate_ledger_report(
        self, data: SeparateLedgerResult, output_path: str
    ) -> Result[str, Exception]:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "律師分帳報表"

            # Headers
            headers = ["日期", "摘要", "部門", "借方金額", "貸方金額", "律師代碼"]
            ws.append(headers)

            # Style Headers
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Data Rows
            for row in data.rows:
                ws.append(
                    [
                        row.date,
                        row.abstract,
                        row.department,
                        row.debit,
                        row.credit,
                        row.lawyer_code,
                    ]
                )

            # Totals Row
            # Assuming data rows are 2 to N+1
            last_row = len(data.rows) + 2
            ws.cell(row=last_row, column=1, value="總計").font = header_font
            ws.cell(row=last_row, column=4, value=data.total_debit).font = header_font
            ws.cell(row=last_row, column=5, value=data.total_credit).font = header_font

            # Save
            wb.save(output_path)
            return Result.success(output_path)

        except Exception as e:
            return Result.failure(
                InfrastructureError(f"Failed to generate report: {e}")
            )
