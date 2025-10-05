from __future__ import annotations

import logging
import os
from tkinter import filedialog, messagebox
from typing import List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from ttk_app.db.crud.lawyer import fetch_all_lawyers, insert_unique_lawyers
from ttk_app.models.worksheets import SeparateAccountsWorksheet
from ttk_app.utils import check_filename_is_valid
from ttk_app.views import MainView
from ttk_app.views.custom_widgets import LawyerSelectionDialog

from .base import BaseController


class LedgerFormatError(Exception):
    """Raised when the source worksheet format does not meet expectations."""

    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


class MainController(BaseController):
    def __init__(self) -> None:
        super().__init__()

        # Attributes
        self.selected_file_name: Optional[str] = None
        self.output_dir: Optional[str] = None
        self._skip_manual_prompts = False
        # Create Window
        self.window = MainView(self)
        self.window.mainloop()

    
    def open_excel_file(self) -> Optional[str]:
        file_is_not_selected = True
        while file_is_not_selected:
            logging.debug("正在詢問用戶要開啟的檔案...")
            filename = filedialog.askopenfilename()

            if filename: # 確定有選擇檔案
                # 檢查副檔名是否與 Excel 相關
                if os.path.splitext(filename)[1] not in [".xls", ".xlsx"]:
                    messagebox.showwarning("警告", "選擇的檔案不是 .xls 或 .xlsx，請重新選擇！")
                    continue
                else:
                    logging.info(f"用戶選擇檔案: {filename}")
                    self.selected_file_name = filename
                    auto_output_dir = os.path.dirname(filename)
                    self.output_dir = auto_output_dir
                    if self.window is not None:
                        self.window.set_source_file_path(filename)
                        self.window.set_output_dir_path(auto_output_dir)
                    file_is_not_selected = False
                    return filename
            
            else: # 用戶案取消
                logging.debug("用戶取消選擇檔案！")
                break
        return None
    
    def choose_output_dir(self) -> Optional[str]:
        folder_is_not_selected = True
        while folder_is_not_selected:
            logging.debug("正在詢問用戶輸出檔案的資料夾...")
            folder_path = filedialog.askdirectory()

            if folder_path:
                logging.info(f"用戶選擇輸出位置: {folder_path}")
                self.output_dir = folder_path
                if self.window is not None:
                    self.window.set_output_dir_path(folder_path)
                folder_is_not_selected = False
                return folder_path
            else:
                break
        return None
    
    def separate_the_ledger(self) -> None:
        # 先檢查【輸出資料夾】是否有給，若無輸出檔名，預設為【output.xlsx】
        if self.output_dir is None:
            messagebox.showerror("錯誤", "請先選擇要輸出的資料夾！")
            return

        output_file_name = self.window.output_filename.get()
        if output_file_name == "輸入輸出檔名" or output_file_name.replace(" ", "") == "": # 如果移除所有空格為空 => 無名稱
            messagebox.showinfo("說明", "沒有給予輸出檔名，將使用 'output' 作為檔名")
            output_file_name = "output"
        elif check_filename_is_valid(output_file_name):
            messagebox.showerror("錯誤", "輸出檔名包含不合法字元")
            return
        if self.selected_file_name is None:
            messagebox.showerror("錯誤", "請先選擇來源檔案！")
            return

        output_file_name = os.path.join(self.output_dir, output_file_name + ".xlsx")
        logging.info(f"資料將儲存於: {output_file_name}")
            


        logging.info(f"開始處理檔案: {self.selected_file_name}, 執行動作: 'separate_the_ledger()'")
        read_engine = self._get_excel_read_engine(self.selected_file_name)
        xls = pd.ExcelFile(self.selected_file_name, engine=read_engine)
        sheet_names = xls.sheet_names

        target_sheet_name = "明細分類帳(依科目+部門)-0_備註欄加上承辦律師"
        if target_sheet_name not in sheet_names:
            messagebox.showwarning("錯誤", "缺乏名為「明細分類帳(依科目+部門)-0_備註欄加上承辦律師」的工作表(sheet)，請檢查檔案中是否有此工作表")
            logging.warning("缺乏名為「明細分類帳(依科目+部門)-0_備註欄加上承辦律師」的工作表(sheet)，請檢查檔案中是否有此工作表")
            return
        
        origin_sheet: DataFrame = pd.read_excel(self.selected_file_name, sheet_name=target_sheet_name, engine=read_engine)
        origin_rows: List[Tuple[object, ...]] = list(origin_sheet.itertuples(index=False, name=None))
        if not origin_rows:
            messagebox.showerror("錯誤", "工作表沒有資料，請確認檔案內容是否正確。")
            self.window.proccess_hint_text.set("格式錯誤：工作表沒有資料")
            return

        try:
            data, total_debit, total_credit = self._build_ledger_entries(origin_rows)
        except LedgerFormatError as format_err:
            messagebox.showerror("格式錯誤", format_err.message)
            self.window.proccess_hint_text.set(f"格式錯誤：{format_err.message}")
            logging.error("Ledger format error: %s", format_err.message)
            return

        try:
            target_sheet = SeparateAccountsWorksheet(output_file_name, self.window)
            target_sheet.write_data_to_worksheet(data, total_debit=total_debit, total_credit=total_credit) # 寫入分帳
            
            # 取出所有 Lawyer Code，進行遍歷
            lawyer_credits: List[dict[str, int]] = []
            lawyers = fetch_all_lawyers()
            for lawyer in lawyers:
                lawyer_total_credit = 0
                for row in data:
                    if row[5] == lawyer.code:
                        lawyer_total_credit += int(row[4])

                lawyer_credits.append({
                    "lawyer_code": lawyer.code,
                    "lawyer_total_credit": lawyer_total_credit,
                })

            logging.debug("Lawyer credit summary: %s", lawyer_credits)


            # 都沒問題，處理完畢，用 Label 告知使用者
            self.window.proccess_hint_text.set("處理完畢！")
        except LedgerFormatError as err:
            messagebox.showerror("格式錯誤", err.message)
            self.window.proccess_hint_text.set(f"格式錯誤：{err.message}")
            logging.error("寫入資料時發生格式錯誤: %s", err.message)
        except Exception as err:
            messagebox.showerror("寫入資料時發生錯誤", "請聯繫管理員！")
            logging.error(f"寫入資料到 'SeparateAccountsWorksheet' 時發生錯誤... Error: {err}")

    def auto_fill_lawyer_codes(self) -> None:
        if self.selected_file_name is None:
            messagebox.showerror("錯誤", "請先選擇來源檔案！")
            return

        target_sheet_name = "明細分類帳(依科目+部門)-0_備註欄加上承辦律師"

        workbook_path = self._ensure_openpyxl_compatible_workbook()
        read_engine = self._get_excel_read_engine(workbook_path)

        try:
            origin_sheet: DataFrame = pd.read_excel(workbook_path, sheet_name=target_sheet_name, engine=read_engine)
        except ValueError:
            messagebox.showerror("錯誤", "無法在來源檔案中找到目標工作表，請檢查檔案內容。")
            return

        origin_rows: List[Tuple[object, ...]] = list(origin_sheet.itertuples(index=False, name=None))
        if not origin_rows:
            messagebox.showinfo("提醒", "工作表沒有資料，無需填入律師代碼。")
            return

        try:
            header_index, _ = self._locate_header(origin_rows)
        except LedgerFormatError as err:
            messagebox.showerror("格式錯誤", err.message)
            return

        lawyer_records = fetch_all_lawyers()
        known_codes = [record.code for record in lawyer_records]

        workbook = load_workbook(workbook_path)
        if target_sheet_name not in workbook.sheetnames:
            messagebox.showerror("錯誤", "來源檔案缺少目標工作表，請確認檔案內容。")
            return
        worksheet = workbook[target_sheet_name]

        data_rows = origin_rows[header_index + 1 :]
        if not data_rows:
            messagebox.showinfo("提醒", "表頭之後沒有資料列，無需填入律師代碼。")
            return

        updated_row_count = 0
        self._skip_manual_prompts = False
        for offset, row in enumerate(data_rows):
            excel_row_number = header_index + offset + 3  # openpyxl 1-based row index
            display_row_number = excel_row_number
            remark_cell = worksheet.cell(row=excel_row_number, column=10)
            remark_value = str(remark_cell.value).strip() if remark_cell.value is not None else ""
            if remark_value:
                continue

            abstract_value = row[1]
            summary_text = str(abstract_value) if abstract_value is not None else ""
            if summary_text.strip() == "" or summary_text.lower() == "nan":
                continue

            date_value = row[0]
            if pd.isna(date_value) or str(date_value).strip() == "" or str(date_value).lower() == "nan":
                continue

            matched_codes = self._match_codes_in_summary(summary_text, known_codes)

            if not matched_codes:
                if self._skip_manual_prompts:
                    continue
                selected_codes = self._prompt_for_codes(summary_text, display_row_number, known_codes)
                if self._skip_manual_prompts or not selected_codes:
                    continue
                insert_unique_lawyers(selected_codes)
                known_codes = sorted(set(known_codes + selected_codes))
                matched_codes = selected_codes

            unique_codes = self._deduplicate_preserve_order(matched_codes)
            remark_cell.value = " ".join(unique_codes)
            updated_row_count += 1

        workbook.save(workbook_path)

        if updated_row_count == 0:
            self.window.proccess_hint_text.set("沒有需要更新的備註欄位。")
        else:
            self.window.proccess_hint_text.set(f"已更新 {updated_row_count} 筆備註欄位。")
        messagebox.showinfo("完成", f"已更新 {updated_row_count} 筆備註欄位。")

    @staticmethod
    def _row_is_empty(row: Tuple[object, ...]) -> bool:
        return all(pd.isna(cell) or str(cell).strip() == "" for cell in row)

    def _build_ledger_entries(
        self, origin_rows: List[Tuple[object, ...]]
    ) -> Tuple[List[List[object]], int, int]:
        header_index, header_row = self._locate_header(origin_rows)

        expected_columns = {
            0: "日期",
            1: "摘要",
            2: "借方金額",
            3: "貸方金額",
            8: "部門",
            9: "備註",
        }

        mismatched_columns: List[str] = []
        for column_index, expected_value in expected_columns.items():
            raw_value = str(header_row[column_index])
            normalized_actual = self._normalize_text(raw_value)
            normalized_expected = self._normalize_text(expected_value)
            if normalized_expected not in normalized_actual:
                mismatched_columns.append(
                    f"第 {column_index + 1} 欄預期為「{expected_value}」，偵測到「{raw_value.strip() or '空白'}」"
                )

        if mismatched_columns:
            raise LedgerFormatError("；".join(mismatched_columns))

        data_rows = origin_rows[header_index + 1 :]
        if not data_rows:
            raise LedgerFormatError("表頭之後沒有資料列，請確認工作表內容。")

        data: List[List[object]] = []
        total_debit = 0
        total_credit = 0
        for offset, row in enumerate(data_rows, start=1):
            if self._row_is_empty(row):
                continue

            excel_row_number = header_index + offset + 2  # 1-based row number (+1 header row, +1 zero index)
            if len(row) < 10:
                raise LedgerFormatError(f"第 {excel_row_number} 行欄位數不足，預期至少 10 欄。")

            date = row[0]
            if pd.isna(date) or str(date).strip() == "":
                # 視為空白列，繼續尋找下一列
                continue

            abstract = str(row[1]).strip() if not pd.isna(row[1]) else ""
            debit_value = self._coerce_amount(row[2], "借方金額", excel_row_number)
            credit_value = self._coerce_amount(row[3], "貸方金額", excel_row_number)

            department_raw = row[8]
            department = str(department_raw).strip() if not pd.isna(department_raw) else ""
            if department == "" or department.lower() == "nan":
                raise LedgerFormatError(f"第 {excel_row_number} 行「部門」欄位為空白。")

            remark = str(row[9]).strip()
            if remark == "" or remark == "nan":
                raise LedgerFormatError(f"第 {excel_row_number} 行「備註」欄位缺少律師代碼。")

            lawyer_code_list = [item.strip() for item in remark.split(" ") if item.strip()]
            if not lawyer_code_list:
                raise LedgerFormatError(f"第 {excel_row_number} 行「備註」欄位缺少律師代碼。")

            insert_unique_lawyers(lawyer_code_list)

            for code in lawyer_code_list:
                separate_debit = int(round(debit_value / len(lawyer_code_list)))
                separate_credit = int(round(credit_value / len(lawyer_code_list)))
                total_debit += separate_debit
                total_credit += separate_credit

                data.append([date, abstract, department, separate_debit, separate_credit, code])

        if not data:
            raise LedgerFormatError("未能從資料中取得任何有效的律師分帳紀錄，請確認來源資料。")

        return data, total_debit, total_credit

    def _locate_header(self, origin_rows: List[Tuple[object, ...]]) -> Tuple[int, Tuple[object, ...]]:
        for index, row in enumerate(origin_rows):
            if self._row_is_empty(row):
                continue

            if len(row) < 10:
                raise LedgerFormatError(f"第 {index + 1} 行欄位數不足，預期至少 10 欄。")

            cell_value = self._normalize_text(row[9])
            if cell_value == self._normalize_text("備註"):
                return index, row

        raise LedgerFormatError("找不到包含「備註」欄位的表頭，請確認來源檔案格式。")

    @staticmethod
    def _coerce_amount(raw_value: object, column_name: str, row_number: int) -> float:
        if pd.isna(raw_value):
            return 0.0

        value_to_convert = raw_value
        if isinstance(raw_value, str):
            value_to_convert = raw_value.replace(",", "").strip()
            if value_to_convert == "":
                return 0.0

        try:
            return float(value_to_convert)
        except (TypeError, ValueError):
            raise LedgerFormatError(
                f"第 {row_number} 行「{column_name}」欄位不是有效的數字：{raw_value}"
            ) from None

    @staticmethod
    def _normalize_text(value: object) -> str:
        text = str(value).strip()
        return text.replace(" ", "").replace("\u3000", "")

    @staticmethod
    def _deduplicate_preserve_order(codes: Sequence[str]) -> List[str]:
        seen = set()
        ordered_codes: List[str] = []
        for code in codes:
            normalized = code.strip()
            if not normalized:
                continue
            if normalized not in seen:
                seen.add(normalized)
                ordered_codes.append(normalized)
        return ordered_codes

    @staticmethod
    def _match_codes_in_summary(summary: str, available_codes: Sequence[str]) -> List[str]:
        normalized_summary = summary
        matched: List[str] = []
        for code in available_codes:
            if code and code in normalized_summary:
                matched.append(code)
        return matched

    def _prompt_for_codes(self, summary: str, row_number: int, available_codes: Sequence[str]) -> Optional[List[str]]:
        if self.window is None:
            return None

        dialog = LawyerSelectionDialog(self.window, summary, row_number, list(available_codes))
        result = dialog.show()
        if getattr(dialog, "skip_remaining", False):
            self._skip_manual_prompts = True
        return result

    @staticmethod
    def _get_excel_read_engine(file_path: str) -> str:
        if file_path.lower().endswith(".xls"):
            return "xlrd"
        return "openpyxl"

    def _ensure_openpyxl_compatible_workbook(self) -> str:
        if self.selected_file_name is None:
            raise LedgerFormatError("尚未選擇來源檔案。")

        file_path = self.selected_file_name
        if not file_path.lower().endswith(".xls"):
            return file_path

        converted_path = os.path.splitext(file_path)[0] + "_converted.xlsx"
        if not os.path.exists(converted_path):
            read_engine = self._get_excel_read_engine(file_path)
            xls = pd.ExcelFile(file_path, engine=read_engine)
            with pd.ExcelWriter(converted_path, engine="openpyxl") as writer:
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        self.selected_file_name = converted_path
        if self.window is not None:
            messagebox.showinfo(
                "資訊",
                f"偵測到 .xls 檔案，已轉換為 {os.path.basename(converted_path)} 以便後續處理。",
                parent=self.window,
            )
        return converted_path
