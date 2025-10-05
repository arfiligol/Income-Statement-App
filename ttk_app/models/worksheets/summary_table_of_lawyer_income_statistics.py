from __future__ import annotations

from typing import List, Optional, Sequence, TYPE_CHECKING

from openpyxl.styles import Alignment, Border, Font, Side
from tkinter import messagebox

if TYPE_CHECKING:
    from ttk_app.views import MainView

from .base import BaseWorksheet


class SummaryTableOfLawyerIncomeStatisticsWorksheet(BaseWorksheet):
    def __init__(self, workbook_path: str, tk_root: Optional["MainView"] = None) -> None:
        self.headers: List[str] = ["承辦律師", "部門", "借方合計", "貸方合計", "備註"]
        super().__init__(workbook_path, "run後檔案再一次統計_律師收入統計總表", tk_root)

    def format_worksheet(self) -> None:
        letters = ["A", "B", "C", "D", "E"]
        width_offset = 2
        self.ws.column_dimensions["A"].width = 15.9 + width_offset
        self.ws.column_dimensions["B"].width = 50.5 + width_offset
        self.ws.column_dimensions["C"].width = 13 + width_offset
        self.ws.column_dimensions["D"].width = 15.9 + width_offset
        self.ws.column_dimensions["E"].width = 15.9 + width_offset

        for col, letter in enumerate(letters, start=1):
            self.ws.merge_cells(f"{letter}1:{letter}2")
            cell = self.ws.cell(row=1, column=col)
            cell.value = self.headers[col - 1]
            cell.font = Font(name="微軟正黑體", size=13)
            for row in self.ws[f"{letters[0]}1:{letters[-1]}2"]:
                for single_cell in row:
                    single_cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_data_to_worksheet(self, data_rows: Sequence[Sequence[object]], **_: object) -> None:
        first_empty_row = self.ws.max_row + 1
        if first_empty_row != 3:
            overwrite_old_data = messagebox.askyesno(
                "偵測到工作表已有資料", "偵測到工作表'run後檔案再一次統計_律師收入統計總表'已有資料，是否覆寫？"
            )
            if overwrite_old_data:
                first_empty_row = 3

        for row_index, row in enumerate(data_rows, start=0):
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=first_empty_row + row_index, column=column_index + 1)
                cell.value = row[column_index]
                cell.font = Font(name="Courier New", size=13)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if column_index in (2, 3):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="bottom", wrap_text=True)

        self.save(self.workbook_path)
