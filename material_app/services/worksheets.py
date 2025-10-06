from __future__ import annotations

from typing import Optional, Sequence

from PySide6.QtWidgets import QMessageBox, QWidget

from ttk_app.models.worksheets.separate_account import SeparateAccountsWorksheet


class QtSeparateAccountsWorksheet(SeparateAccountsWorksheet):
    """SeparateAccountsWorksheet that uses Qt message boxes instead of Tk dialogs."""

    def __init__(self, parent: QWidget, workbook_path: str) -> None:
        self._qt_parent = parent
        super().__init__(workbook_path, tk_root=None)

    def init_ws(self) -> None:
        if self.is_new_workbook and "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        if self.sheet_title in self.wb.sheetnames:
            response = QMessageBox.question(
                self._qt_parent,
                "工作表存在",
                f"工作表'{self.sheet_title}' 已經存在，你想要覆蓋它嗎？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if response == QMessageBox.StandardButton.Yes:
                self.wb.remove(self.wb[self.sheet_title])
                self.create_worksheet()
            else:
                self.ws = self.wb[self.sheet_title]
        else:
            self.create_worksheet()

    def write_data_to_worksheet(
        self,
        data_rows: Sequence[Sequence[object]],
        *,
        total_debit: Optional[int] = None,
        total_credit: Optional[int] = None,
    ) -> None:
        first_empty_row = self.ws.max_row + 1
        if first_empty_row != 3:
            response = QMessageBox.question(
                self._qt_parent,
                "問題",
                "偵測到工作表'程式RUN後檔案'已有資料，是否覆寫？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if response == QMessageBox.StandardButton.Yes:
                first_empty_row = 3

        for row_index, row in enumerate(data_rows):
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=first_empty_row + row_index, column=column_index + 1)
                cell.value = row[column_index]
                self._style_data_cell(column_index, cell)

        if total_credit is not None or total_debit is not None:
            self.ws.append(["", "合計", "", total_debit, total_credit, ""])
            total_row = self.ws.max_row
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=total_row, column=column_index + 1)
                self._style_data_cell(column_index, cell)

        self.save(self.workbook_path)

    def _style_data_cell(self, column_index: int, cell) -> None:
        from openpyxl.styles import Alignment, Border, Font, Side

        cell.font = Font(name="Courier New", size=13)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        if column_index == 4:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(vertical="bottom", wrap_text=True)
