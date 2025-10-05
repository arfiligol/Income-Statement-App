from __future__ import annotations

import logging
import os
from typing import List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame
from PySide6.QtCore import QObject, Slot
from PySide6.QtWidgets import QFileDialog, QMessageBox

from material_app.dialogs import LawyerSelectionDialog
from material_app.models.dto import AppState
from material_app.views.main_window import MainWindow
from ttk_app.controllers.main_controller import LedgerFormatError
from ttk_app.db.crud.lawyer import fetch_all_lawyers, insert_unique_lawyers
from ttk_app.models.worksheets import SeparateAccountsWorksheet
from ttk_app.utils import check_filename_is_valid


class WorkflowController(QObject):
    """Handles workflow-tab user interactions."""

    def __init__(self, window: MainWindow, state: AppState) -> None:
        super().__init__(window)
        self.window = window
        self.state = state
        self._skip_manual_prompts = False

        self.window.selectSourceRequested.connect(self.select_source_file)
        self.window.selectOutputDirRequested.connect(self.select_output_dir)
        self.window.actionSelected.connect(self.on_action_selected)
        self.window.submitRequested.connect(self.on_submit)

    # ------------------------------------------------------------------
    @Slot()
    def select_source_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self.window,
            "選擇來源檔案",
            filter="Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            logging.debug("使用者取消選擇來源檔案")
            return

        if os.path.splitext(file_path)[1].lower() not in {".xlsx", ".xls"}:
            QMessageBox.warning(self.window, "警告", "選擇的檔案不是 .xls 或 .xlsx，請重新選擇！")
            return

        logging.info("用戶選擇檔案: %s", file_path)
        self.state.source_path = file_path
        self.window.set_source_path(file_path)

        auto_output_dir = os.path.dirname(file_path)
        self.state.output_dir = auto_output_dir
        self.window.set_output_dir(auto_output_dir)
        self.window.set_status_message("已載入來源檔案")
        self.window.set_submit_state("warning")

    @Slot()
    def select_output_dir(self) -> None:
        dir_path = QFileDialog.getExistingDirectory(self.window, "選擇輸出資料夾")
        if not dir_path:
            logging.debug("使用者取消選擇輸出資料夾")
            return

        logging.info("用戶選擇輸出位置: %s", dir_path)
        self.state.output_dir = dir_path
        self.window.set_output_dir(dir_path)
        self.window.set_status_message("已更新輸出資料夾")
        self.window.set_submit_state("warning")

    @Slot(str)
    def on_action_selected(self, action_name: str) -> None:
        if not action_name:
            return
        self.state.selected_action = action_name
        self.window.set_selected_action(action_name)
        self.window.set_status_message(f"已選擇功能：{self._action_label(action_name)}")
        self.window.set_submit_state("warning")

    @Slot()
    def on_submit(self) -> None:
        source_path = self.state.source_path
        if source_path is None:
            QMessageBox.warning(self.window, "警告", "請先選擇來源檔案！")
            self.window.set_submit_state("danger")
            return

        if self.state.output_dir is None:
            auto_dir = os.path.dirname(source_path)
            self.state.output_dir = auto_dir
            self.window.set_output_dir(auto_dir)

        filename = self.window.get_output_filename().strip()
        if filename in {"", "輸入輸出檔名"}:
            QMessageBox.warning(self.window, "警告", "請先輸入輸出檔案名稱！")
            self.window.set_status_message("請輸入輸出檔案名稱")
            self.window.set_submit_state("danger")
            return
        if check_filename_is_valid(filename):
            QMessageBox.critical(self.window, "錯誤", "輸出檔名包含不合法字元")
            self.window.set_submit_state("danger")
            return

        self.state.output_filename = filename

        action = self.state.selected_action
        if action == "auto_fill_remark":
            self.auto_fill_lawyer_codes()
        elif action == "separate_the_ledger":
            self.separate_the_ledger()
        else:
            QMessageBox.information(self.window, "提示", "請先選擇要執行的功能！")
            self.window.set_status_message("請先選擇要執行的功能！")
            self.window.set_submit_state("warning")

    # ------------------------------------------------------------------
    def auto_fill_lawyer_codes(self) -> None:
        source_path = self.state.source_path
        if source_path is None:
            return

        workbook_path = self._ensure_openpyxl_compatible_workbook(source_path)
        read_engine = self._get_excel_read_engine(workbook_path)

        target_sheet_name = "明細分類帳(依科目+部門)-0_備註欄加上承辦律師"
        try:
            origin_sheet: DataFrame = pd.read_excel(workbook_path, sheet_name=target_sheet_name, engine=read_engine)
        except ValueError:
            QMessageBox.critical(self.window, "錯誤", "無法在來源檔案中找到目標工作表，請檢查檔案內容。")
            self.window.set_submit_state("danger")
            return

        origin_rows: List[Tuple[object, ...]] = list(origin_sheet.itertuples(index=False, name=None))
        if not origin_rows:
            QMessageBox.information(self.window, "提示", "工作表沒有資料，無需填入律師代碼。")
            self.window.set_status_message("工作表沒有資料")
            self.window.set_submit_state("warning")
            return

        try:
            header_index, _ = self._locate_header(origin_rows)
        except LedgerFormatError as err:
            QMessageBox.critical(self.window, "格式錯誤", err.message)
            self.window.set_status_message(err.message)
            self.window.set_submit_state("danger")
            return

        lawyer_records = fetch_all_lawyers()
        known_codes = [record.code for record in lawyer_records]

        workbook = load_workbook(workbook_path)
        if target_sheet_name not in workbook.sheetnames:
            QMessageBox.critical(self.window, "錯誤", "來源檔案缺少目標工作表，請確認檔案內容。")
            self.window.set_submit_state("danger")
            return

        worksheet = workbook[target_sheet_name]
        data_rows = origin_rows[header_index + 1 :]
        if not data_rows:
            QMessageBox.information(self.window, "提示", "表頭之後沒有資料列，無需填入律師代碼。")
            self.window.set_status_message("無資料需要更新")
            self.window.set_submit_state("warning")
            return

        updated_count = 0
        self._skip_manual_prompts = False

        for offset, row in enumerate(data_rows):
            if self._skip_manual_prompts:
                break

            excel_row_number = header_index + offset + 3
            display_row_number = excel_row_number
            remark_cell = worksheet.cell(row=excel_row_number, column=10)
            remark_value = str(remark_cell.value).strip() if remark_cell.value is not None else ""
            if remark_value:
                continue

            date_value = row[0]
            if pd.isna(date_value) or str(date_value).strip() in {"", "nan"}:
                continue

            summary_text = str(row[1]) if row[1] is not None else ""
            if summary_text.strip() == "" or summary_text.lower() == "nan":
                continue

            matched_codes = self._match_codes_in_summary(summary_text, known_codes)

            if not matched_codes:
                selected_codes = self._prompt_for_codes(summary_text, display_row_number, known_codes)
                if self._skip_manual_prompts:
                    break
                if not selected_codes:
                    continue
                insert_unique_lawyers(selected_codes)
                known_codes = sorted(set(known_codes + selected_codes))
                matched_codes = selected_codes

            remark_cell.value = " ".join(self._deduplicate_preserve_order(matched_codes))
            updated_count += 1

        workbook.save(workbook_path)

        self.window.set_status_message(f"已更新 {updated_count} 筆備註欄位。")
        self.window.set_submit_state("success")

    def separate_the_ledger(self) -> None:
        output_dir = self.state.output_dir
        source_path = self.state.source_path
        output_name = self.state.output_filename
        if output_dir is None or source_path is None or output_name is None:
            return

        read_engine = self._get_excel_read_engine(source_path)
        xls = pd.ExcelFile(source_path, engine=read_engine)
        sheet_names = xls.sheet_names

        target_sheet_name = "明細分類帳(依科目+部門)-0_備註欄加上承辦律師"
        if target_sheet_name not in sheet_names:
            QMessageBox.warning(self.window, "錯誤", "缺乏名為「明細分類帳(依科目+部門)-0_備註欄加上承辦律師」的工作表(sheet)")
            logging.warning("缺乏目標工作表")
            self.window.set_submit_state("danger")
            return

        origin_sheet: DataFrame = pd.read_excel(source_path, sheet_name=target_sheet_name, engine=read_engine)
        origin_array = origin_sheet.values
        if len(origin_array[0]) != 10:
            QMessageBox.critical(
                self.window,
                "錯誤",
                f"格式與當初給定的不一樣，原本偵測到十個欄位，現在偵測到: {len(origin_array[0])} 個欄位",
            )
            self.window.set_submit_state("danger")
            return

        data: List[List[object]] = []
        total_debit = 0
        total_credit = 0
        ready_for_data = False
        for row in origin_array:
            if row[9] != "備註" and not ready_for_data:
                continue
            if row[9] == "備註":
                ready_for_data = True
                continue

            date = row[0]
            if pd.isna(date):
                continue
            abstract = row[1]
            debit = row[2]
            credit = row[3]
            department = row[8]
            remark = row[9]

            lawyer_code_list = str(remark).split(" ")
            filtered_code_list = [item for item in lawyer_code_list if item]
            if not filtered_code_list:
                continue
            insert_unique_lawyers(filtered_code_list)

            if pd.isna(debit):
                debit = 0
            if pd.isna(credit):
                credit = 0

            for code in filtered_code_list:
                separate_debit = int(round(float(debit) / len(filtered_code_list)))
                separate_credit = int(round(float(credit) / len(filtered_code_list)))
                total_debit += separate_debit
                total_credit += separate_credit
                data.append([date, abstract, department, separate_debit, separate_credit, code])

        try:
            output_path = os.path.join(output_dir, f"{output_name}.xlsx")
            target_sheet = SeparateAccountsWorksheet(output_path, None)
            target_sheet.write_data_to_worksheet(data, total_debit=total_debit, total_credit=total_credit)
        except Exception as err:  # pylint: disable=broad-except
            QMessageBox.critical(self.window, "寫入資料時發生錯誤", "請聯繫管理員！")
            logging.error("寫入目標工作表時發生錯誤: %s", err)
            self.window.set_submit_state("danger")
            return

        self.window.set_status_message("處理完畢！")
        self.window.set_submit_state("success")

    # ------------------------------------------------------------------
    def _prompt_for_codes(self, summary: str, row_number: int, available_codes: Sequence[str]) -> Optional[List[str]]:
        dialog = LawyerSelectionDialog(summary, row_number, list(available_codes), parent=self.window)
        if dialog.exec():
            if dialog.skip_remaining:
                self._skip_manual_prompts = True
            return dialog.selected_codes
        return None

    @staticmethod
    def _get_excel_read_engine(file_path: str) -> str:
        return "xlrd" if file_path.lower().endswith(".xls") else "openpyxl"

    def _ensure_openpyxl_compatible_workbook(self, file_path: str) -> str:
        if not file_path.lower().endswith(".xls"):
            return file_path

        converted_path = os.path.splitext(file_path)[0] + "_converted.xlsx"
        if not os.path.exists(converted_path):
            read_engine = self._get_excel_read_engine(file_path)
            xls = pd.ExcelFile(file_path, engine=read_engine)
            with pd.ExcelWriter(converted_path, engine="openpyxl") as writer:
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        return converted_path

    @staticmethod
    def _match_codes_in_summary(summary: str, available_codes: Sequence[str]) -> List[str]:
        return [code for code in available_codes if code and code in summary]

    @staticmethod
    def _deduplicate_preserve_order(codes: Sequence[str]) -> List[str]:
        seen = set()
        ordered: List[str] = []
        for code in codes:
            clean = code.strip()
            if clean and clean not in seen:
                seen.add(clean)
                ordered.append(clean)
        return ordered

    def _locate_header(self, origin_rows: List[Tuple[object, ...]]) -> Tuple[int, Tuple[object, ...]]:
        for index, row in enumerate(origin_rows):
            if all(pd.isna(cell) or str(cell).strip() == "" for cell in row):
                continue
            if len(row) < 10:
                raise LedgerFormatError(f"第 {index + 1} 行欄位數不足，預期至少 10 欄。")
            cell_value = str(row[9]).strip()
            if cell_value == "備註":
                return index, row
        raise LedgerFormatError("找不到包含「備註」欄位的表頭，請確認來源檔案格式。")

    @staticmethod
    def _action_label(action_name: str) -> str:
        return {
            "auto_fill_remark": "摘要抓律師代碼",
            "separate_the_ledger": "律師收入明細",
        }.get(action_name, action_name)
