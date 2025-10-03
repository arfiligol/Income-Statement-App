from __future__ import annotations

from tkinter import messagebox
from typing import TYPE_CHECKING, List, Optional, Sequence

from openpyxl.styles import Alignment, Border, Font, Side

if TYPE_CHECKING:
    from views import MainView

from .base import BaseWorksheet


class SeparateAccountsWorksheet(BaseWorksheet):
    def __init__(self, workbook_path: str, tk_root: Optional["MainView"] = None) -> None:
        self.headers: List[str] = ["日期", "摘要", "部門", "借方金額", "貸方金額", "承辦律師"]
        super().__init__(workbook_path, "程式RUN後檔案", tk_root)

    # 用於格式化、設定工作表樣貌
    def format_worksheet(self) -> None:
        letters = ["A", "B", "C", "D", "E", "F"]
        # 設定欄寬
        width_offset = 2
        self.ws.column_dimensions["A"].width = 15.9 + width_offset
        self.ws.column_dimensions["B"].width = 50.5 + width_offset
        self.ws.column_dimensions["C"].width = 13 + width_offset
        self.ws.column_dimensions["D"].width = 15.9 + width_offset
        self.ws.column_dimensions["E"].width = 15.9 + width_offset
        self.ws.column_dimensions["F"].width = 15.9 + width_offset


        # 建立欄位
        for col, letter in enumerate(letters, start=1):
            self.ws.merge_cells(f"{letter}1:{letter}2")
            cell = self.ws.cell(row=1, column=col)

            # 設定 Cell
            cell.value = self.headers[col - 1]  # 給值
            cell.font = Font(name="微軟正黑體", size=13)
            for row in self.ws[f"{letters[0]}1:{letters[-1]}2"]:
                for single_cell in row:
                    single_cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    
    def write_data_to_worksheet(
        self,
        data_rows: Sequence[Sequence[object]],
        *,
        total_debit: Optional[int] = None,
        total_credit: Optional[int] = None,
    ) -> None:
        first_empty_row = self.ws.max_row + 1
        # 檢測是否有資料
        if first_empty_row != 3:
            overwrite_old_data = messagebox.askyesno("問題", "偵測到工作表'程式RUN後檔案'已有資料，是否覆寫？")
            if overwrite_old_data:
                # 若決定覆寫，將數據寫入第一行設在 row = 2
                first_empty_row = 3
        
        # 開始寫入資料到工作表
        for row_index, row in enumerate(data_rows, start=0):  # 因為 first_empty_row 會設為最新空格，因此 row_index 從 0 開始
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=first_empty_row + row_index, column=column_index + 1)
                cell.value = row[column_index]
                
                cell.font = Font(name="Courier New", size=13)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if column_index == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        
        if total_credit is not None or total_debit is not None:
            self.ws.append(["", "合計", "", total_debit, total_credit, ""])
            total_row = self.ws.max_row
            for column_index, _ in enumerate(self.headers):
                cell = self.ws.cell(row=total_row, column=column_index + 1)
                
                cell.font = Font(name="Courier New", size=13)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                if column_index == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="bottom", wrap_text=True)
        self.save(self.workbook_path)
