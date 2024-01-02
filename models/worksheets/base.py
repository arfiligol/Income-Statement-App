import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import tkinter as tk
from tkinter import messagebox

class BaseWorksheet:
    def __init__(self, workbook_path, sheet_title, tk_root = None):
        self.workbook_path = workbook_path
        self.sheet_title = sheet_title
        self.root = tk_root

        if not os.path.exists(workbook_path):
            self.wb = Workbook()
            self.is_new_workbook = True
        else:
            self.is_new_workbook = False
            self.wb = load_workbook(workbook_path)

        self.init_ws()

    def init_ws(self):
        # 如果是新建 Workbook 而有預設建立的工作表，刪除它
        if self.is_new_workbook:
            self.wb.remove(self.wb["Sheet"])

        # 如果目標工作表存在則使用它，否則創建一個新的工作表
        if self.sheet_title in self.wb.sheetnames:
            # 確認是否覆蓋掉已有的工作表
            if self.root: # 檢查是否有 tk 視窗，沒有的話改用 console 詢問
                response = messagebox.askyesno("工作表存在", f"工作表'{self.sheet_title}' 已經存在，你想要覆蓋它嗎？")
                if response:
                    # 刪除創建新的
                    self.wb.remove(self.wb[self.sheet_title])
                    self.create_worksheet()
                else:
                    # 使用已有的工作表
                    self.ws = self.wb[self.sheet_title]
            else:
                response = input(f"工作表 '{self.sheet_title}' 已存在，你想要覆蓋它嗎？(Yes/No, Default: No, Only 'Yes' acceptable.)")
                if response == "Yes":
                    # 刪除創建新的
                    self.wb.remove(self.wb[self.sheet_title])
                    self.create_worksheet()
                else:
                    # 使用已有的工作表
                    self.ws = self.wb[self.sheet_title]

        else: # 沒有該工作表，直接創建一個新的
            self.create_worksheet()

    def create_worksheet(self):
        self.ws = self.wb.create_sheet(self.sheet_title)
        self.format_worksheet()
        self.save(self.workbook_path)

    def save(self, filename):
        try:
            self.wb.save(filename)
    
        except PermissionError as err:
            if self.root:
                messagebox.showerror(f"無法儲存變更到檔案名稱{filename}", f"請確認目標檔案'{filename}'是否已經關閉，若透過其他程式開啟該檔案，則無法在此程式更改該檔案！")
            else:
                print(f"無法儲存變更到檔案名稱{filename}", f"請確認目標檔案'{filename}'是否已經關閉，若透過其他程式開啟該檔案，則無法在此程式更改該檔案！")
                print(err)

    def format_worksheet(self):
        # 省略原有 format_worksheet 內容
        pass

    def write_data_to_worksheet(self, data_rows):
        # 省略原有 write_data_to_worksheet 內容
        pass